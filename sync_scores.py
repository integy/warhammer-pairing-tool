#!/usr/bin/env python3
"""
WTC Pairing Matrix Sync Script
Syncs scores from Google Sheets → team JSON files in warhammer-pairing-tool.

Usage:
  python3 sync_scores.py                           # sync from Google Sheets
  python3 sync_scores.py --file matrix.xlsx        # sync from local xlsx file
  python3 sync_scores.py --dry-run                 # preview changes only
"""

from __future__ import annotations
import sys, os, json, re, argparse
import urllib.request
from typing import Optional

SHEET_ID = "15jeILF17cdh0fhWAXN-SI4WVF3TsM667B1Y_-Hb3LUE"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
TEAMS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "public", "teams")

# HK player name mapping: Google Sheet row label → JSON name
# We extract the name before " - " and map it
HK_NAME_MAP = {
    "q b": "Quentin",
    "kristian kammerer": "Kris",
    "terence tsui": "Terrence",
    "gabriel leung": "Gabrielle",
    "ivan wong": "Ivan",
    "kevin lee": "Kevin",
    "jireh mok": "Jireh",
    "eric 'corazon' yu": "Eric",
    "eric yu": "Eric",
    "eric": "Eric",
}


def team_to_key(name: str) -> str:
    """Map sheet/team name to JSON key."""
    return name.strip().lower().replace(" ", "-").replace("/", "-")


def extract_hk_name(row_label: str) -> Optional[str]:
    """Extract HK player name from row label like 'q b - Chaos Space Marines'."""
    name_part = row_label.split(" - ")[0].strip().lower()
    # Direct match
    if name_part in HK_NAME_MAP:
        return HK_NAME_MAP[name_part]
    # Fuzzy match
    for key, val in HK_NAME_MAP.items():
        if key in name_part or name_part in key:
            return val
    print(f"  ⚠ Unknown HK player: '{name_part}' (from '{row_label}')")
    return None


def extract_opponent_name(col_header: str) -> str:
    """Extract opponent name from column header like 'Dave Kerr\\nBlood Angels'."""
    return col_header.split("\n")[0].strip()


def sync_from_xlsx(filepath: str, dry_run: bool = False):
    """Sync scores from an xlsx file."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"📂 Loaded: {filepath} ({len(wb.sheetnames)} sheets)")

    updates = []  # [(team_key, player_name, scores_dict)]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        team_key = team_to_key(sheet_name)

        # Check if team JSON exists
        team_path = os.path.join(TEAMS_DIR, f"{team_key}-team.json")
        if not os.path.exists(team_path):
            print(f"  ⚠ {sheet_name}: no team JSON found (key={team_key}) — skip")
            continue

        # Read column headers (row 1, cols C-J) → opponent player names
        opponent_names = []
        for c in range(3, 11):
            val = ws.cell(row=1, column=c).value
            if val:
                opponent_names.append(extract_opponent_name(str(val)))

        if not opponent_names:
            print(f"  ⚠ {sheet_name}: no opponent headers — skip")
            continue

        # Read HK player names (rows 3-10, col A)
        hk_names = []
        for r in range(3, 11):
            val = ws.cell(row=r, column=1).value
            if val:
                hk_name = extract_hk_name(str(val))
                hk_names.append(hk_name)
            else:
                hk_names.append(None)

        # Read scores matrix
        scores_by_opponent = {}  # {opponent_name: {hk_name: score}}

        for ci, opp_name in enumerate(opponent_names):
            col = ci + 3  # 1-indexed column
            scores = {}
            for ri, hk_name in enumerate(hk_names):
                row = ri + 3
                if hk_name is None:
                    continue
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        score = float(cell.value)
                        # Convert whole numbers to int for cleaner JSON
                        if score == int(score):
                            score = int(score)
                        scores[hk_name] = score
                    except (ValueError, TypeError):
                        pass
            if scores:
                scores_by_opponent[opp_name] = scores

        if not scores_by_opponent:
            continue  # No scores in this sheet

        # Update team JSON
        with open(team_path) as f:
            team_data = json.load(f)

        # Resolve opponent header names to actual JSON player names.
        # Headers are usually "Name\nArmy", but sometimes "Name Army" (space-separated,
        # no newline) or other variants, so do exact match first then prefix/substring.
        player_names = [p["name"] for p in team_data["players"]]
        resolved = {}
        for opp_name, scores in scores_by_opponent.items():
            match = None
            if opp_name in player_names:
                match = opp_name
            else:
                for pn in player_names:
                    if opp_name.startswith(pn) or pn in opp_name:
                        match = pn
                        break
            if match:
                resolved[match] = scores
            else:
                print(f"  ⚠ {sheet_name}: cannot resolve opponent '{opp_name}' — skip")
        scores_by_opponent = resolved

        changed = False
        for p in team_data["players"]:
            if p["name"] in scores_by_opponent:
                new_scores = scores_by_opponent[p["name"]]
                old_scores = p.get("scores", {})
                if new_scores != old_scores:
                    p["scores"] = new_scores
                    changed = True
                    updates.append((team_key, p["name"], new_scores))

        if changed and not dry_run:
            with open(team_path, "w") as f:
                json.dump(team_data, f, indent=2)

        if scores_by_opponent:
            total = sum(len(s) for s in scores_by_opponent.values())
            print(f"  ✅ {sheet_name}: {total} scores ({len(scores_by_opponent)} players)")

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Total: {len(updates)} player score updates")
    if dry_run and updates:
        for team_key, name, scores in updates:
            score_str = ", ".join(f"{k}={v}" for k, v in scores.items())
            print(f"  {team_key}/{name}: {score_str}")

    return updates


def main():
    parser = argparse.ArgumentParser(description="Sync WTC pairing scores from Google Sheets")
    parser.add_argument("--file", "-f", help="Local xlsx file path (instead of Google Sheets)")
    parser.add_argument("--dry-run", "-n", action="store_true", help="Preview changes without writing")
    args = parser.parse_args()

    if args.file:
        filepath = args.file
    else:
        print(f"⬇ Downloading from Google Sheets...")
        filepath = "/tmp/wtc_sync_matrix.xlsx"
        urllib.request.urlretrieve(SHEET_URL, filepath)
        print(f"   Saved to {filepath}")

    sync_from_xlsx(filepath, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
